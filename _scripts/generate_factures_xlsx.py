"""Génère le tableau maître Factures-fournisseurs.xlsx pour MDS.
Structure :
  - Onglet README : workflow + légende
  - Onglet Fournisseurs : annuaire (1 ligne par fournisseur)
  - Onglet <Fournisseur> : factures + virements + récap pour ce fournisseur
"""
from pathlib import Path
from datetime import date
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

OUT = Path(r"G:\Mon Drive\Factures-fournisseurs\Factures-fournisseurs.xlsx")

# --- Styles ---
TITLE = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
TITLE_FILL = PatternFill("solid", fgColor="1F4E78")
H = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
H_FILL = PatternFill("solid", fgColor="4472C4")
SUB = Font(name="Calibri", size=11, bold=True)
SUB_FILL = PatternFill("solid", fgColor="D9E2F3")
TOTAL_FILL = PatternFill("solid", fgColor="FFE699")
BORDER = Border(*[Side(style="thin", color="BFBFBF")] * 4)
CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
RIGHT = Alignment(horizontal="right", vertical="center")

wb = Workbook()

# =========================
# ONGLET README
# =========================
ws = wb.active
ws.title = "README"
ws.column_dimensions["A"].width = 4
ws.column_dimensions["B"].width = 30
ws.column_dimensions["C"].width = 90

ws["B2"] = "Tableau factures fournisseurs MDS"
ws["B2"].font = Font(size=16, bold=True, color="1F4E78")

readme = [
    ("", ""),
    ("Objectif", "Centraliser toutes les factures et paiements fournisseurs, 1 onglet par fournisseur."),
    ("", ""),
    ("Workflow Gmail → tableau", ""),
    ("1.", "Reçois une facture par mail (Gmail metallier.mds@gmail.com)"),
    ("2.", "Tu mets le label 📥 Factures à traiter (clic droit ou règle auto par expéditeur)"),
    ("3.", "Apps Script Gmail (cron 1 h) → dépose le PDF dans G:\\Mon Drive\\Factures-fournisseurs\\inbox\\"),
    ("4.", "Tu demandes : « Claude, traite l'inbox »"),
    ("5.", "Claude analyse chaque PDF, ajoute une ligne dans l'onglet du fournisseur, déplace le PDF dans archive/<fournisseur>/"),
    ("", ""),
    ("Structure d'un onglet fournisseur", ""),
    ("Bloc 1", "Fiche identité (code client, SIRET, contact, IBAN, conditions paiement)"),
    ("Bloc 2", "Tableau des factures (1 ligne par facture, statut, lien PDF)"),
    ("Bloc 3", "Tableau des virements/paiements"),
    ("Bloc 4", "Récap : total facturé, total payé, reste dû"),
    ("", ""),
    ("Statuts factures", "À régler / Acompte versé / Soldée / Litige"),
    ("", ""),
    ("Maintenance", "Tableau modifié manuellement OU par Claude à la demande. Sync auto via Google Drive."),
]
for i, (a, b) in enumerate(readme, start=4):
    ws.cell(row=i, column=2, value=a).font = SUB if a and not a[0].isdigit() else Font()
    ws.cell(row=i, column=3, value=b).alignment = LEFT

# =========================
# ONGLET FOURNISSEURS (annuaire)
# =========================
fws = wb.create_sheet("Fournisseurs")
headers_f = [
    "Code", "Raison sociale", "SIRET / TVA", "Adresse", "Code client MDS",
    "Contact principal", "Téléphone", "Email",
    "Banque", "IBAN", "BIC",
    "Conditions paiement", "Notes"
]
for j, h in enumerate(headers_f, start=1):
    c = fws.cell(row=1, column=j, value=h)
    c.font = H; c.fill = H_FILL; c.alignment = CENTER; c.border = BORDER

# Première fiche : Gantois
row_gantois = [
    "GAN", "GANTOIS INDUSTRIES SAS",
    "TVA FR 38 531 918 456 / RCS Epinal 531 918 456",
    "B.P. 307 - 88105 ST-DIE-DES-VOSGES CEDEX",
    "C015634",
    "Jean-Pierre DEPREAUX", "06 72 95 90 93", "jean-pierre.depreaux@gantois.com",
    "CA Normandie", "FR76 1660 6533 5000 1585 3989 776", "AGRIFRPP866",
    "Acompte 30% commande, solde livraison, virement",
    "Pénalité stockage 65€HT/sem/volée si retard livraison"
]
for j, v in enumerate(row_gantois, start=1):
    c = fws.cell(row=2, column=j, value=v); c.border = BORDER; c.alignment = LEFT

widths_f = [8, 28, 35, 35, 14, 22, 16, 32, 16, 30, 14, 35, 35]
for j, w in enumerate(widths_f, start=1):
    fws.column_dimensions[get_column_letter(j)].width = w
fws.row_dimensions[2].height = 45
fws.freeze_panes = "A2"

# =========================
# ONGLET GANTOIS
# =========================
gws = wb.create_sheet("Gantois")

def block_title(ws, row, text, span=10):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    c = ws.cell(row=row, column=1, value=text)
    c.font = TITLE; c.fill = TITLE_FILL; c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[row].height = 22

# --- Bloc 1 : Fiche identité ---
block_title(gws, 1, "FICHE FOURNISSEUR — GANTOIS INDUSTRIES SAS")
fiche = [
    ("Code interne",    "GAN"),
    ("Code client MDS", "C015634"),
    ("SIRET / TVA",     "TVA FR 38 531 918 456 — RCS Epinal 531 918 456"),
    ("Adresse",         "B.P. 307 — 88105 ST-DIE-DES-VOSGES CEDEX"),
    ("Tél siège",       "+33 (0)3 29 55 21 43"),
    ("Contact",         "Jean-Pierre DEPREAUX"),
    ("Téléphone",       "06 72 95 90 93"),
    ("Email",           "jean-pierre.depreaux@gantois.com"),
    ("Banque",          "CA Normandie"),
    ("IBAN",            "FR76 1660 6533 5000 1585 3989 776"),
    ("BIC / SWIFT",     "AGRIFRPP866"),
    ("Conditions",      "Acompte 30% à la commande, solde à livraison, virement bancaire"),
    ("Pénalités",       "Stockage 65 € HT/semaine/volée au-delà de la date de livraison convenue"),
]
for i, (k, v) in enumerate(fiche, start=2):
    a = gws.cell(row=i, column=1, value=k); a.font = SUB; a.fill = SUB_FILL; a.border = BORDER; a.alignment = LEFT
    b = gws.cell(row=i, column=2, value=v); b.border = BORDER; b.alignment = LEFT
    gws.merge_cells(start_row=i, start_column=2, end_row=i, end_column=10)

# --- Bloc 2 : Factures ---
start = 2 + len(fiche) + 1
block_title(gws, start, "FACTURES")
headers_fac = ["Date", "N° facture", "N° commande", "Réf MDS", "Désignation",
               "Total HT", "TVA", "Total TTC", "Échéance", "Statut", "PDF / Notes"]
for j, h in enumerate(headers_fac, start=1):
    c = gws.cell(row=start+1, column=j, value=h)
    c.font = H; c.fill = H_FILL; c.alignment = CENTER; c.border = BORDER

# Facture du 19/05/2026
fac_row = start + 2
gantois_facture = [
    date(2026, 5, 19),
    "FACPTE000001247 (proforma)",
    "CV266462",
    "CCF 000685",
    "Escalier intérieur",
    None,            # HT non visible sur page 2/2
    None,            # TVA non visible
    11244.00,
    date(2026, 5, 19),
    "Acompte versé",
    "Page 1/2 manquante (détail articles). Pénalité stockage 65€HT/sem/volée si retard."
]
for j, v in enumerate(gantois_facture, start=1):
    c = gws.cell(row=fac_row, column=j, value=v); c.border = BORDER; c.alignment = LEFT
    if isinstance(v, float):
        c.number_format = '#,##0.00 "€"'; c.alignment = RIGHT
    if isinstance(v, date):
        c.number_format = "dd/mm/yyyy"

# --- Bloc 3 : Virements ---
start2 = fac_row + 2
block_title(gws, start2, "VIREMENTS / PAIEMENTS")
headers_vir = ["Date virement", "Type", "Montant", "Mode", "Facture liée", "Référence virement", "Notes"]
for j, h in enumerate(headers_vir, start=1):
    c = gws.cell(row=start2+1, column=j, value=h)
    c.font = H; c.fill = H_FILL; c.alignment = CENTER; c.border = BORDER

vir_row = start2 + 2
vir = [
    date(2026, 5, 19),
    "Acompte 30%",
    3373.20,
    "Virement bancaire",
    "FACPTE000001247 / CCF 000685",
    "À renseigner (libellé virement banque)",
    "Acompte commande escalier intérieur CV266462"
]
for j, v in enumerate(vir, start=1):
    c = gws.cell(row=vir_row, column=j, value=v); c.border = BORDER; c.alignment = LEFT
    if isinstance(v, float):
        c.number_format = '#,##0.00 "€"'; c.alignment = RIGHT
    if isinstance(v, date):
        c.number_format = "dd/mm/yyyy"

# --- Bloc 4 : Récap ---
start3 = vir_row + 2
block_title(gws, start3, "RÉCAP")
recap = [
    ("Total facturé TTC", "=H{}".format(fac_row), '#,##0.00 "€"'),
    ("Total payé",        "=C{}".format(vir_row), '#,##0.00 "€"'),
    ("Reste dû",          "=H{}-C{}".format(fac_row, vir_row), '#,##0.00 "€"'),
]
for i, (lbl, formula, fmt) in enumerate(recap, start=start3+1):
    a = gws.cell(row=i, column=1, value=lbl); a.font = SUB; a.fill = TOTAL_FILL; a.border = BORDER
    b = gws.cell(row=i, column=2, value=formula); b.font = Font(bold=True); b.fill = TOTAL_FILL
    b.border = BORDER; b.number_format = fmt; b.alignment = RIGHT
    gws.merge_cells(start_row=i, start_column=2, end_row=i, end_column=3)

# Largeurs colonnes Gantois
widths_g = [18, 28, 16, 14, 30, 12, 10, 14, 14, 18, 40]
for j, w in enumerate(widths_g, start=1):
    gws.column_dimensions[get_column_letter(j)].width = w
gws.freeze_panes = "A2"

# Sauvegarde
OUT.parent.mkdir(parents=True, exist_ok=True)
wb.save(OUT)
print(f"OK -> {OUT}")
print(f"Taille : {OUT.stat().st_size} octets")
